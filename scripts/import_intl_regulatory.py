#!/usr/bin/env python3
"""
Import international regulatory data into the ChemIP terminology.db.

Step 1: EPA Consolidated List of Lists (from downloaded XLSX)
Step 2: ECHA Annex VI CLP (from downloaded XLSX)
Step 3: International Conventions (Rotterdam, Stockholm, Montreal) - hardcoded
Step 4: Match to chemical_terms and report statistics
"""

import os
import re
import sqlite3
import sys

DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'terminology.db')
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')


def normalize_cas(cas_str):
    """Normalize CAS number: strip whitespace, remove leading zeros from segments."""
    if not cas_str:
        return None
    cas_str = str(cas_str).strip()
    # Must match pattern like 123-45-6
    m = re.match(r'^0*(\d+)-(\d+)-(\d+)$', cas_str)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # Try without hyphens but with correct digit pattern
    return cas_str


def create_table(conn):
    """Create intl_regulatory table with indexes."""
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS intl_regulatory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cas_no TEXT,
            regulation TEXT NOT NULL,
            detail TEXT,
            source TEXT NOT NULL,
            chemical_name TEXT
        )
    ''')
    # Drop old indexes if they exist (for re-runs)
    cursor.execute("DROP INDEX IF EXISTS idx_intl_reg_cas")
    cursor.execute("DROP INDEX IF EXISTS idx_intl_reg_regulation")
    cursor.execute("DROP INDEX IF EXISTS idx_intl_reg_source")

    cursor.execute("CREATE INDEX idx_intl_reg_cas ON intl_regulatory(cas_no)")
    cursor.execute("CREATE INDEX idx_intl_reg_regulation ON intl_regulatory(regulation)")
    cursor.execute("CREATE INDEX idx_intl_reg_source ON intl_regulatory(source)")
    conn.commit()
    print("[OK] Created intl_regulatory table with indexes")


def clear_table(conn):
    """Clear existing data for clean re-import."""
    conn.execute("DELETE FROM intl_regulatory")
    conn.commit()
    print("[OK] Cleared existing intl_regulatory data")


# ============================================================
# STEP 1: EPA Consolidated List of Lists
# ============================================================
def import_epa_list(conn):
    """Parse EPA Consolidated List of Lists XLSX and insert into intl_regulatory."""
    import openpyxl

    xlsx_path = os.path.join(DATA_DIR, 'epa_consolidated_list.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"[SKIP] EPA file not found: {xlsx_path}")
        return 0

    print(f"[...] Loading EPA file: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['List of Lists']

    # Column mapping (0-indexed from row 9):
    # 0: Name, 1: CAS Number, 3: CAA 112(r)(7) TQ, 4: CERCLA HS RQ,
    # 5: CWA 311 HS TQ, 6: EPCRA 302 EHS TPQ, 7: EPCRA 304 EHS RQ, 8: EPCRA 313 TRI
    regulation_cols = {
        3: ('CAA_112R', 'Threshold Quantity'),
        4: ('CERCLA', 'Reportable Quantity'),
        5: ('CWA_311', 'Threshold Quantity'),
        6: ('EPCRA_302', 'Threshold Planning Quantity'),
        7: ('EPCRA_304', 'Reportable Quantity'),
        8: ('EPCRA_313', 'TRI'),
    }

    rows_to_insert = []
    row_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 10:  # Skip header rows (data starts at row 10, 0-indexed)
            continue

        name = row[0]
        cas_raw = row[1]
        if not name or not cas_raw:
            continue

        cas = normalize_cas(str(cas_raw).strip())
        # Skip category codes (e.g., N171, N583) - not real CAS numbers
        if cas and not re.match(r'^\d+-\d+-\d+$', cas):
            # Some entries have category codes like "N171" instead of CAS
            cas = None

        for col_idx, (reg_name, detail_prefix) in regulation_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip() not in ('', 'None'):
                detail_str = str(val).strip()
                if detail_str:
                    detail = f"{detail_prefix}: {detail_str}" if detail_prefix != 'TRI' else detail_str
                    rows_to_insert.append((cas, reg_name, detail, 'EPA', str(name).strip()))
                    row_count += 1

    cursor = conn.cursor()
    cursor.executemany(
        "INSERT INTO intl_regulatory (cas_no, regulation, detail, source, chemical_name) VALUES (?, ?, ?, ?, ?)",
        rows_to_insert
    )
    conn.commit()
    wb.close()
    print(f"[OK] EPA: Inserted {row_count} regulatory entries")
    return row_count


# ============================================================
# STEP 2: ECHA Annex VI CLP
# ============================================================
def import_echa_clp(conn):
    """Parse ECHA Annex VI CLP XLSX and insert into intl_regulatory."""
    import openpyxl

    xlsx_path = os.path.join(DATA_DIR, 'echa_annex_vi_clp.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"[SKIP] ECHA CLP file not found: {xlsx_path}")
        print("       Attempting download...")
        if not download_echa_clp(xlsx_path):
            return 0

    print(f"[...] Loading ECHA CLP file: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # Try to find the right sheet
    target_sheet = None
    for name in wb.sheetnames:
        if 'annex' in name.lower() or 'table' in name.lower() or 'ATP' in name:
            target_sheet = name
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    print(f"  Using sheet: {target_sheet}")
    ws = wb[target_sheet]

    # Find header row and column positions
    # Known structure: Row 4 has main headers, Row 5 has sub-headers, data from row 6
    header_row_idx = None
    cas_col = 3       # default for ATP21
    name_col = 1
    hazard_class_col = 4
    hazard_stmt_col = 5

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 20:
            break
        if row is None:
            continue
        row_strs = [str(c).lower() if c else '' for c in row]
        for j, cell in enumerate(row_strs):
            if 'cas' in cell and ('no' in cell or 'number' in cell):
                cas_col = j
                # Data starts 2 rows after header (skip sub-header row)
                header_row_idx = i + 1
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        header_row_idx = 5  # fallback
        print("  [WARN] Could not find CAS header, using fallback")

    print(f"  CAS col={cas_col}, name col={name_col}, data starts after row {header_row_idx}")

    rows_to_insert = []
    row_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue

        cas_raw = row[cas_col] if cas_col < len(row) else None
        if not cas_raw:
            continue

        cas_str = str(cas_raw).strip()
        # Some entries have multiple CAS separated by newlines or spaces
        cas_candidates = re.findall(r'\d+-\d+-\d+', cas_str)
        if not cas_candidates:
            continue

        chem_name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ''
        chem_name = chem_name.replace('\n', ' ').strip()

        # Build detail from hazard class + statement
        details = []
        if hazard_class_col < len(row) and row[hazard_class_col]:
            details.append(str(row[hazard_class_col]).strip().replace('\n', '; '))
        if hazard_stmt_col < len(row) and row[hazard_stmt_col]:
            details.append(str(row[hazard_stmt_col]).strip().replace('\n', '; '))
        detail = ' | '.join(details) if details else 'Harmonised Classification'

        for cas in cas_candidates:
            cas = normalize_cas(cas)
            if cas:
                rows_to_insert.append((cas, 'EU_CLP', detail, 'ECHA', chem_name))
                row_count += 1

    cursor = conn.cursor()
    cursor.executemany(
        "INSERT INTO intl_regulatory (cas_no, regulation, detail, source, chemical_name) VALUES (?, ?, ?, ?, ?)",
        rows_to_insert
    )
    conn.commit()
    wb.close()
    print(f"[OK] ECHA CLP: Inserted {row_count} regulatory entries")
    return row_count


def download_echa_clp(target_path):
    """Try to download ECHA CLP Excel. Returns True if successful."""
    try:
        import urllib.request
        # ECHA provides the file at a known URL pattern
        urls = [
            "https://echa.europa.eu/documents/10162/17218/annex_vi_clp_table_atp21_en.xlsx",
            "https://echa.europa.eu/documents/10162/17218/annex_vi_clp_table_atp20_en.xlsx",
            "https://echa.europa.eu/documents/10162/17218/annex_vi_clp_table_atp19_en.xlsx",
        ]
        for url in urls:
            try:
                print(f"  Trying: {url}")
                req = urllib.request.Request(url, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                })
                with urllib.request.urlopen(req, timeout=30) as resp:
                    data = resp.read()
                    if len(data) > 10000:  # Sanity check
                        with open(target_path, 'wb') as f:
                            f.write(data)
                        print(f"  Downloaded {len(data)} bytes")
                        return True
            except Exception as e:
                print(f"  Failed: {e}")
                continue
        return False
    except Exception as e:
        print(f"  Download failed: {e}")
        return False


# ============================================================
# STEP 3: International Conventions (Hardcoded)
# ============================================================

# Rotterdam Convention Annex III chemicals (57 chemicals)
ROTTERDAM_CHEMICALS = [
    ("2104-64-5", "EPN"),
    ("1646-88-4", "Alachlor (formulations with >7% active ingredient)"),
    ("15972-60-8", "Alachlor"),
    ("309-00-2", "Aldrin"),
    ("116-06-3", "Aldicarb (formulations with >15% active ingredient)"),
    ("1397-94-0", "Asbestos (Actinolite)"),
    ("12172-73-5", "Asbestos (Amosite)"),
    ("77536-66-4", "Asbestos (Anthophyllite)"),
    ("12001-29-5", "Asbestos (Chrysotile)"),
    ("12001-28-4", "Asbestos (Crocidolite)"),
    ("12172-67-7", "Asbestos (Tremolite)"),
    ("1332-21-4", "Asbestos (all forms)"),
    ("86-50-0", "Azinphos-methyl"),
    ("17804-35-2", "Benomyl (formulations)"),
    ("82657-04-3", "Bifenthrin"),
    ("1689-84-5", "Bromoxynil heptanoate"),
    ("55285-14-8", "Carbofuran (formulations)"),
    ("1563-66-2", "Carbofuran"),
    ("55285-14-8", "Carbosulfan"),
    ("143-50-0", "Chlordecone"),
    ("57-74-9", "Chlordane"),
    ("6164-98-3", "Chlordimeform"),
    ("2921-88-2", "Chlorpyrifos (formulations)"),
    ("76-44-8", "Heptachlor"),
    ("118-74-1", "Hexachlorobenzene"),
    ("87-68-3", "Hexachlorobutadiene"),
    ("58-89-9", "Lindane (HCH gamma isomer)"),
    ("608-73-1", "Hexachlorocyclohexane (mixed isomers)"),
    ("72-43-5", "Methoxychlor"),
    ("298-00-0", "Parathion-methyl (formulations)"),
    ("56-38-2", "Parathion"),
    ("85-44-9", "Phthalic anhydride"),
    ("1336-36-3", "Polychlorinated biphenyls (PCB)"),
    ("32534-81-9", "Pentabromodiphenyl ether"),
    ("36483-60-0", "Hexabromodiphenyl ether"),
    ("1031078-04-6", "Decabromodiphenyl ether"),
    ("40487-42-1", "Pendimethalin"),
    ("82-68-8", "Pentachloronitrobenzene (PCNB)"),
    ("87-86-5", "Pentachlorophenol and its salts"),
    ("7439-92-1", "Lead compounds"),
    ("62-73-7", "Dichlorvos (formulations)"),
    ("94-75-7", "2,4-D (formulations)"),
    ("131-52-2", "Sodium pentachlorophenate"),
    ("2439-01-2", "Chinomethionat"),
    ("115-32-2", "Dicofol"),
    ("50-29-3", "DDT"),
    ("60-57-1", "Dieldrin"),
    ("330-54-1", "Diuron (formulations)"),
    ("72-20-8", "Endrin"),
    ("106-93-4", "Ethylene dibromide (EDB)"),
    ("75-21-8", "Ethylene oxide"),
    ("2385-85-5", "Mirex"),
    ("6923-22-4", "Monocrotophos"),
    ("1582-09-8", "Trifluralin"),
    ("900-95-8", "Tributyltin compounds"),
    ("76-87-9", "Fentin hydroxide"),
    ("8001-35-2", "Toxaphene/Camphechlor"),
]

# Stockholm Convention POPs (all annexes, ~37 chemicals/groups)
STOCKHOLM_CHEMICALS = [
    ("309-00-2", "Aldrin", "Annex A - Elimination"),
    ("57-74-9", "Chlordane", "Annex A - Elimination"),
    ("143-50-0", "Chlordecone", "Annex A - Elimination"),
    ("2385-85-5", "Mirex", "Annex A - Elimination"),
    ("50-29-3", "DDT", "Annex B - Restriction"),
    ("60-57-1", "Dieldrin", "Annex A - Elimination"),
    ("72-20-8", "Endrin", "Annex A - Elimination"),
    ("76-44-8", "Heptachlor", "Annex A - Elimination"),
    ("118-74-1", "Hexachlorobenzene", "Annex A/C - Elimination"),
    ("319-84-6", "Alpha-hexachlorocyclohexane", "Annex A - Elimination"),
    ("319-85-7", "Beta-hexachlorocyclohexane", "Annex A - Elimination"),
    ("58-89-9", "Lindane (gamma-HCH)", "Annex A - Elimination"),
    ("608-93-5", "Pentachlorobenzene", "Annex A/C - Elimination"),
    ("8001-35-2", "Toxaphene", "Annex A - Elimination"),
    ("1336-36-3", "Polychlorinated biphenyls (PCB)", "Annex A/C - Elimination"),
    ("1763-23-1", "Perfluorooctane sulfonic acid (PFOS)", "Annex B - Restriction"),
    ("335-67-1", "Perfluorooctanoic acid (PFOA)", "Annex A - Elimination"),
    ("355-46-4", "Perfluorohexane sulfonic acid (PFHxS)", "Annex A - Elimination"),
    ("36355-01-8", "Hexabromobiphenyl", "Annex A - Elimination"),
    ("25637-99-4", "Hexabromocyclododecane (HBCD)", "Annex A - Elimination"),
    ("32534-81-9", "Pentabromodiphenyl ether (pentaBDE)", "Annex A - Elimination"),
    ("36483-60-0", "Hexabromodiphenyl ether (hexaBDE)", "Annex A - Elimination"),
    ("68631-49-2", "Decabromodiphenyl ether (decaBDE)", "Annex A - Elimination"),
    ("5436-43-1", "Tetrabromodiphenyl ether (tetraBDE)", "Annex A - Elimination"),
    ("60348-60-9", "Pentabromodiphenyl ether isomer", "Annex A - Elimination"),
    ("87-68-3", "Hexachlorobutadiene", "Annex A/C - Elimination"),
    ("608-73-1", "Hexachlorocyclohexane (mixed isomers)", "Annex A - Elimination"),
    ("319-86-8", "Delta-hexachlorocyclohexane", "Annex A - Elimination"),
    ("56-23-5", "Carbon tetrachloride (short-chain chlorinated paraffin precursor)", "Annex A - Elimination"),
    ("85535-84-8", "Short-chain chlorinated paraffins (SCCPs)", "Annex A - Elimination"),
    ("2921-88-2", "Chlorpyrifos", "Annex A - Elimination"),
    ("115-32-2", "Dicofol", "Annex A - Elimination"),
    ("2227-13-6", "Tetrachlorvinfos", "Annex A - Elimination (proposed)"),
    ("82-68-8", "Pentachloronitrobenzene (PCNB)", "Annex A - Elimination"),
    ("87-86-5", "Pentachlorophenol", "Annex A - Elimination"),
    ("107-06-2", "Ethylene dichloride (EDC) - unintentional", "Annex C - Unintentional"),
    ("127-18-4", "Perchloroethylene - unintentional", "Annex C - Unintentional"),
    ("67-72-1", "Hexachloroethane", "Annex C - Unintentional (proposed)"),
    ("307-35-7", "Perfluorooctane sulfonyl fluoride (PFOSF)", "Annex B - Restriction"),
    ("1024-57-3", "Heptachlor epoxide", "Annex A - Elimination"),
    ("39001-02-0", "Octachlorodibenzofuran (PCDF marker)", "Annex C - Unintentional"),
]

# Montreal Protocol - Ozone Depleting Substances
MONTREAL_CHEMICALS = [
    # Class I - CFCs (Annex A, Group I)
    ("75-69-4", "CFC-11 (Trichlorofluoromethane)", "Class I - CFC, ODP=1.0"),
    ("75-71-8", "CFC-12 (Dichlorodifluoromethane)", "Class I - CFC, ODP=1.0"),
    ("76-13-1", "CFC-113 (1,1,2-Trichlorotrifluoroethane)", "Class I - CFC, ODP=0.8"),
    ("76-14-2", "CFC-114 (Dichlorotetrafluoroethane)", "Class I - CFC, ODP=1.0"),
    ("76-15-3", "CFC-115 (Monochloropentafluoroethane)", "Class I - CFC, ODP=0.6"),
    # Halons (Annex A, Group II)
    ("353-59-3", "Halon 1211 (Bromochlorodifluoromethane)", "Class I - Halon, ODP=3.0"),
    ("75-63-8", "Halon 1301 (Bromotrifluoromethane)", "Class I - Halon, ODP=10.0"),
    ("124-73-2", "Halon 2402 (Dibromotetrafluoroethane)", "Class I - Halon, ODP=6.0"),
    # Other CFCs (Annex B, Group I)
    ("75-72-9", "CFC-13 (Chlorotrifluoromethane)", "Class I - CFC, ODP=1.0"),
    ("354-56-3", "CFC-111 (Pentachlorofluoroethane)", "Class I - CFC, ODP=1.0"),
    ("76-12-0", "CFC-112 (Tetrachlorodifluoroethane)", "Class I - CFC, ODP=1.0"),
    ("422-78-6", "CFC-211 (Heptachlorofluoropropane)", "Class I - CFC, ODP=1.0"),
    ("3182-26-1", "CFC-212 (Hexachlorodifluoropropane)", "Class I - CFC, ODP=1.0"),
    ("2354-06-5", "CFC-213 (Pentachlorotrifluoropropane)", "Class I - CFC, ODP=1.0"),
    ("29255-31-0", "CFC-214 (Tetrachlorotetrafluoropropane)", "Class I - CFC, ODP=1.0"),
    ("4259-43-2", "CFC-215 (Trichloropentafluoropropane)", "Class I - CFC, ODP=1.0"),
    ("661-97-2", "CFC-216 (Dichlorohexafluoropropane)", "Class I - CFC, ODP=1.0"),
    ("422-86-6", "CFC-217 (Chloroheptafluoropropane)", "Class I - CFC, ODP=1.0"),
    # Carbon tetrachloride (Annex B, Group II)
    ("56-23-5", "Carbon tetrachloride", "Class I - ODP=1.1"),
    # Methyl chloroform (Annex B, Group III)
    ("71-55-6", "Methyl chloroform (1,1,1-Trichloroethane)", "Class I - ODP=0.1"),
    # Methyl bromide (Annex E)
    ("74-83-9", "Methyl bromide", "Class I - ODP=0.6"),
    # HBFCs (Annex C, Group II) - selected
    ("151-67-7", "Halothane (HBFC-2B1)", "Class I - HBFC"),
    # Bromochloromethane (Annex C, Group III)
    ("74-97-5", "Bromochloromethane", "Class I - ODP=0.12"),
    # Class II - HCFCs (Annex C, Group I)
    ("75-43-4", "HCFC-21 (Dichlorofluoromethane)", "Class II - HCFC, ODP=0.04"),
    ("75-45-6", "HCFC-22 (Chlorodifluoromethane)", "Class II - HCFC, ODP=0.055"),
    ("593-70-4", "HCFC-31 (Chlorofluoromethane)", "Class II - HCFC, ODP=0.02"),
    ("354-14-3", "HCFC-121 (Tetrachlorofluoroethane)", "Class II - HCFC, ODP=0.01-0.04"),
    ("354-21-2", "HCFC-122 (Trichlorodifluoroethane)", "Class II - HCFC, ODP=0.02-0.08"),
    ("306-83-2", "HCFC-123 (Dichlorotrifluoroethane)", "Class II - HCFC, ODP=0.02"),
    ("2837-89-0", "HCFC-124 (Chlorotetrafluoroethane)", "Class II - HCFC, ODP=0.022"),
    ("359-28-4", "HCFC-131 (Trichlorofluoroethane)", "Class II - HCFC, ODP=0.007-0.05"),
    ("1649-08-7", "HCFC-132b (Dichlorodifluoroethane)", "Class II - HCFC, ODP=0.008-0.05"),
    ("75-88-7", "HCFC-133a (Chlorotrifluoroethane)", "Class II - HCFC, ODP=0.02-0.06"),
    ("1717-00-6", "HCFC-141b (Dichlorofluoroethane)", "Class II - HCFC, ODP=0.11"),
    ("75-68-3", "HCFC-142b (Chlorodifluoroethane)", "Class II - HCFC, ODP=0.065"),
    ("422-26-4", "HCFC-221 (Hexachlorofluoropropane)", "Class II - HCFC, ODP=0.015-0.07"),
    ("422-49-1", "HCFC-222 (Pentachlorodifluoropropane)", "Class II - HCFC, ODP=0.01-0.09"),
    ("422-52-6", "HCFC-223 (Tetrachlorotrifluoropropane)", "Class II - HCFC, ODP=0.01-0.08"),
    ("422-54-8", "HCFC-224 (Trichlorotetrafluoropropane)", "Class II - HCFC, ODP=0.01-0.09"),
    ("422-56-0", "HCFC-225ca (Dichloropentafluoropropane)", "Class II - HCFC, ODP=0.025"),
    ("507-55-1", "HCFC-225cb (Dichloropentafluoropropane)", "Class II - HCFC, ODP=0.033"),
    ("431-87-8", "HCFC-226 (Chlorohexafluoropropane)", "Class II - HCFC, ODP=0.02-0.10"),
    ("421-94-3", "HCFC-231 (Pentachlorofluoropropane)", "Class II - HCFC, ODP=0.05-0.09"),
    ("460-89-9", "HCFC-232 (Tetrachlorodifluoropropane)", "Class II - HCFC, ODP=0.008-0.10"),
    ("7125-84-0", "HCFC-233 (Trichlorotrifluoropropane)", "Class II - HCFC, ODP=0.007-0.23"),
    ("425-94-5", "HCFC-234 (Dichlorotetrafluoropropane)", "Class II - HCFC, ODP=0.01-0.28"),
    ("460-92-4", "HCFC-235 (Chloropentafluoropropane)", "Class II - HCFC, ODP=0.03-0.52"),
    ("666-27-3", "HCFC-241 (Tetrachlorofluoropropane)", "Class II - HCFC, ODP=0.004-0.09"),
    ("460-63-9", "HCFC-242 (Trichlorodifluoropropane)", "Class II - HCFC, ODP=0.005-0.13"),
    ("460-69-5", "HCFC-243 (Dichlorotrifluoropropane)", "Class II - HCFC, ODP=0.007-0.12"),
    ("421-41-0", "HCFC-244 (Chlorotetrafluoropropane)", "Class II - HCFC, ODP=0.009-0.14"),
    ("819-00-1", "HCFC-251 (Trichlorofluoropropane)", "Class II - HCFC, ODP=0.001-0.01"),
    ("420-97-3", "HCFC-252 (Dichlorodifluoropropane)", "Class II - HCFC, ODP=0.005-0.04"),
    ("460-35-5", "HCFC-253 (Chlorotrifluoropropane)", "Class II - HCFC, ODP=0.003-0.03"),
    ("420-46-2", "HCFC-261 (Dichlorofluoropropane)", "Class II - HCFC, ODP=0.002-0.02"),
    ("430-55-7", "HCFC-262 (Chlorodifluoropropane)", "Class II - HCFC, ODP=0.002-0.02"),
    ("460-13-9", "HCFC-271 (Chlorofluoropropane)", "Class II - HCFC, ODP=0.001-0.03"),
]


def import_conventions(conn):
    """Insert hardcoded international convention data."""
    cursor = conn.cursor()
    count = 0

    # Rotterdam Convention
    for cas, name in ROTTERDAM_CHEMICALS:
        cursor.execute(
            "INSERT INTO intl_regulatory (cas_no, regulation, detail, source, chemical_name) VALUES (?, ?, ?, ?, ?)",
            (normalize_cas(cas), 'ROTTERDAM_PIC', 'Annex III - Prior Informed Consent', 'Rotterdam Convention', name)
        )
        count += 1
    print(f"[OK] Rotterdam Convention: {count} entries")

    # Stockholm Convention
    stk_count = 0
    for cas, name, annex in STOCKHOLM_CHEMICALS:
        cursor.execute(
            "INSERT INTO intl_regulatory (cas_no, regulation, detail, source, chemical_name) VALUES (?, ?, ?, ?, ?)",
            (normalize_cas(cas), 'STOCKHOLM_POP', annex, 'Stockholm Convention', name)
        )
        stk_count += 1
    count += stk_count
    print(f"[OK] Stockholm Convention: {stk_count} entries")

    # Montreal Protocol
    mtl_count = 0
    for cas, name, detail in MONTREAL_CHEMICALS:
        cursor.execute(
            "INSERT INTO intl_regulatory (cas_no, regulation, detail, source, chemical_name) VALUES (?, ?, ?, ?, ?)",
            (normalize_cas(cas), 'MONTREAL_ODS', detail, 'Montreal Protocol', name)
        )
        mtl_count += 1
    count += mtl_count
    print(f"[OK] Montreal Protocol: {mtl_count} entries")

    conn.commit()
    return count


# ============================================================
# STEP 4: Match to chemical_terms
# ============================================================
def report_matches(conn):
    """Report how many of our 117K chemicals matched to each regulation."""
    cursor = conn.cursor()

    print("\n" + "=" * 70)
    print("MATCHING REPORT: intl_regulatory vs chemical_terms")
    print("=" * 70)

    # Total chemical_terms
    cursor.execute("SELECT COUNT(*) FROM chemical_terms")
    total_chems = cursor.fetchone()[0]
    print(f"\nTotal chemicals in chemical_terms: {total_chems:,}")

    # Total intl_regulatory entries
    cursor.execute("SELECT COUNT(*) FROM intl_regulatory")
    total_reg = cursor.fetchone()[0]
    print(f"Total entries in intl_regulatory: {total_reg:,}")

    # Distinct CAS numbers in intl_regulatory
    cursor.execute("SELECT COUNT(DISTINCT cas_no) FROM intl_regulatory WHERE cas_no IS NOT NULL")
    distinct_cas = cursor.fetchone()[0]
    print(f"Distinct CAS numbers in intl_regulatory: {distinct_cas:,}")

    # We need to handle CAS normalization for matching
    # chemical_terms may have leading-zero CAS like '024012-08-6'
    # Let's do a flexible match

    print("\n--- Matches by regulation ---")
    cursor.execute("SELECT DISTINCT regulation FROM intl_regulatory ORDER BY regulation")
    regulations = [row[0] for row in cursor.fetchall()]

    for reg in regulations:
        # Count distinct CAS in this regulation
        cursor.execute(
            "SELECT COUNT(DISTINCT cas_no) FROM intl_regulatory WHERE regulation=? AND cas_no IS NOT NULL",
            (reg,)
        )
        reg_cas_count = cursor.fetchone()[0]

        # Match against chemical_terms (exact match)
        cursor.execute("""
            SELECT COUNT(DISTINCT ct.cas_no)
            FROM chemical_terms ct
            INNER JOIN intl_regulatory ir ON ct.cas_no = ir.cas_no
            WHERE ir.regulation = ? AND ct.cas_no IS NOT NULL AND ct.cas_no != '' AND ct.cas_no != '-'
        """, (reg,))
        exact_match = cursor.fetchone()[0]

        print(f"  {reg:20s}: {reg_cas_count:5,} CAS in regulation, {exact_match:5,} matched to chemical_terms")

    # Overall match
    cursor.execute("""
        SELECT COUNT(DISTINCT ct.cas_no)
        FROM chemical_terms ct
        INNER JOIN intl_regulatory ir ON ct.cas_no = ir.cas_no
        WHERE ct.cas_no IS NOT NULL AND ct.cas_no != '' AND ct.cas_no != '-'
    """)
    overall_match = cursor.fetchone()[0]
    print(f"\n  {'TOTAL (any reg)':20s}: {distinct_cas:5,} CAS in regulations, {overall_match:5,} matched to chemical_terms")
    print(f"  Match rate: {overall_match / distinct_cas * 100:.1f}% of regulatory CAS found in our DB")
    print(f"  Coverage:   {overall_match / total_chems * 100:.2f}% of our chemicals have regulatory data")

    # Source breakdown
    print("\n--- Entries by source ---")
    cursor.execute("SELECT source, COUNT(*) FROM intl_regulatory GROUP BY source ORDER BY COUNT(*) DESC")
    for source, cnt in cursor.fetchall():
        print(f"  {source:25s}: {cnt:,} entries")


def main():
    db_path = os.path.abspath(DB_PATH)
    print(f"Database: {db_path}")
    print(f"File exists: {os.path.exists(db_path)}")

    conn = sqlite3.connect(db_path, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")

    try:
        # Create table
        create_table(conn)
        clear_table(conn)

        # Step 1: EPA
        print("\n--- STEP 1: EPA Consolidated List of Lists ---")
        epa_count = import_epa_list(conn)

        # Step 2: ECHA CLP
        print("\n--- STEP 2: ECHA Annex VI CLP ---")
        echa_count = import_echa_clp(conn)

        # Step 3: International Conventions
        print("\n--- STEP 3: International Conventions ---")
        conv_count = import_conventions(conn)

        # Step 4: Report
        report_matches(conn)

    finally:
        conn.close()


if __name__ == '__main__':
    main()
